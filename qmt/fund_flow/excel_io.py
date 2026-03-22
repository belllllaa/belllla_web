# -*- coding: utf-8 -*-
"""
写入主工作簿：汇总 sheet（市值亿元、流入万元）+ 每日资金流入汇总 sheet（宽表：股票，1/1主力，1/1散户，1/2主力，1/2散户… 仅交易日）。
"""
from __future__ import annotations

from pathlib import Path
from datetime import datetime
import re

import pandas as pd
from openpyxl import load_workbook


def _date_to_col_prefix(d: str) -> str:
    """'2026-03-09' -> '3/9'，用于列名"""
    if not d:
        return ""
    d = str(d).strip()
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})", d)
    if m:
        return f"{int(m.group(2))}/{int(m.group(3))}"
    m = re.match(r"(\d{4})(\d{2})(\d{2})", d.replace("-", ""))
    if m:
        return f"{int(m.group(2))}/{int(m.group(3))}"
    return d


def _normalize_code_column(ws, col_idx: int = 1) -> None:
    """将指定 sheet 的某一列统一为 6 位文本股票代码，防止 000 开头被 Excel 省略。"""
    from openpyxl.worksheet.worksheet import Worksheet  # type: ignore

    assert isinstance(ws, Worksheet)
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=col_idx)
        v = cell.value
        if v is None:
            continue
        s = str(v).strip().replace(".0", "")
        if s.isdigit():
            cell.value = s.zfill(6)
        else:
            cell.value = s
        cell.number_format = "@"


def write_summary_workbook(summary_df: pd.DataFrame, output_path: str | Path) -> None:
    """
    仅写入汇总 sheet 的工作簿：
    - 汇总：全量指标；市值亿元（流通市值已在上游取整）、流入万元。
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # 准备数据列顺序
    summary = summary_df.copy()
    cols_luobo = [
        "股票代码", "股票名称", "最新价", "涨跌幅",
        "当日净流入(万)", "占成交金额",
        "主力净流入(万)", "散户净流入(万)",
        "所属行业", "流通市值(亿)", "主力-散户净流入(万)",
        "前3日主力净流入之和(亿)", "前5日主力净流入之和(亿)", "前10日主力净流入之和(亿)", "前20日主力净流入之和(亿)", "前40日主力净流入之和(亿)", "前60日主力净流入之和(亿)",
    ]
    for c in cols_luobo:
        if c not in summary.columns:
            summary[c] = None
    summary = summary[[c for c in cols_luobo if c in summary.columns]]
    if "股票代码" in summary.columns:
        summary["股票代码"] = (
            summary["股票代码"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
            .str.zfill(6)
        )

    # 优先使用模板：fund_flow_summary_template.xlsx
    template_path = output_path.parent / "fund_flow_summary_template.xlsx"
    if template_path.exists():
        wb = load_workbook(template_path)
        # 只保留名为“汇总”的 sheet
        if "汇总" in wb.sheetnames:
            ws = wb["汇总"]
        else:
            ws = wb.active
            ws.title = "汇总"
        # 清空除表头外的数据行
        max_row = ws.max_row
        if max_row > 1:
            ws.delete_rows(2, max_row - 1)
        # 将数据按列名写入（保留模板样式）
        header_map = {}
        for col in range(1, ws.max_column + 1):
            name = ws.cell(row=1, column=col).value
            if isinstance(name, str):
                header_map[name.strip()] = col
        for row_idx, (_, row) in enumerate(summary.iterrows(), start=2):
            for col_name, value in row.items():
                col_idx = header_map.get(col_name)
                if col_idx is None:
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
        _normalize_code_column(ws, col_idx=1)
        wb.save(output_path)
        wb.close()
    else:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            summary.to_excel(writer, sheet_name="汇总", index=False)
            wb = writer.book
            if "汇总" in wb.sheetnames:
                _normalize_code_column(wb["汇总"], col_idx=1)


def write_daily_wide_workbook(daily_wide_df: pd.DataFrame, output_path: str | Path) -> None:
    """
    仅写入每日资金流入宽表的工作簿：
    - sheet 名称：每日资金流入汇总
    - 列：股票(6 位文本)、股票名称、各交易日 M/D主力、M/D散户、M/D涨跌幅（主力/散户单位为亿元）
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    daily = daily_wide_df.copy()
    if "股票" in daily.columns:
        daily["股票"] = (
            daily["股票"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
            .str.zfill(6)
        )

    template_path = output_path.parent / "daily_fund_flow_wide_template.xlsx"
    if template_path.exists():
        wb = load_workbook(template_path)
        if "每日资金流入汇总" in wb.sheetnames:
            ws = wb["每日资金流入汇总"]
        else:
            ws = wb.active
            ws.title = "每日资金流入汇总"
        # 清空除表头外的数据行
        max_row = ws.max_row
        if max_row > 1:
            ws.delete_rows(2, max_row - 1)
        # 按列名写入
        from openpyxl.utils.dataframe import dataframe_to_rows

        header_map = {}
        for col in range(1, ws.max_column + 1):
            name = ws.cell(row=1, column=col).value
            if isinstance(name, str):
                header_map[name.strip()] = col
        for row_idx, (_, row) in enumerate(daily.iterrows(), start=2):
            for col_name, value in row.items():
                col_idx = header_map.get(col_name)
                if col_idx is None:
                    continue
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
        _normalize_code_column(ws, col_idx=1)
        wb.save(output_path)
        wb.close()
    else:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            daily.to_excel(writer, sheet_name="每日资金流入汇总", index=False)
            wb = writer.book
            if "每日资金流入汇总" in wb.sheetnames:
                _normalize_code_column(wb["每日资金流入汇总"], col_idx=1)


def append_daily_fund_flow(
    existing_path: str | Path,
    daily_df: pd.DataFrame,
    snapshot_date: str,
) -> None:
    """
    在已有工作簿的「每日资金流入汇总」宽表末尾追加当日列（M/D主力、M/D散户、M/D涨跌幅），主力/散户单位万元。
    若接口中出现新股（不在宽表中），会用 outer 合并自动追加新行，上市前日期列为空。
    daily_df 需含列：代码、主力净流入(万)、散户净流入(万)，可选 涨跌幅；建议含 名称 以便新股有股票名称。
    """
    path = Path(existing_path)
    if not path.exists():
        return
    prefix = _date_to_col_prefix(snapshot_date)
    # 新增当日数据：统一代码为 6 位字符串
    add = daily_df[["代码", "主力净流入", "散户净流入"]].copy()
    add["股票"] = (
        add["代码"]
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
        .str.zfill(6)
    )
    # daily_df 中主力/散户已是万元，直接写入
    add[f"{prefix}主力"] = add["主力净流入"].round(2)
    add[f"{prefix}散户"] = add["散户净流入"].round(2)
    add_cols = ["股票", f"{prefix}主力", f"{prefix}散户"]
    if "涨跌幅" in daily_df.columns:
        add[f"{prefix}涨跌幅"] = daily_df["涨跌幅"].map(
            lambda x: f"{x:.2f}%" if pd.notna(x) else None
        ).values
        add_cols.append(f"{prefix}涨跌幅")
    # 新股名称：用于 outer 合并后补全新行的股票名称
    code_to_name = None
    if "名称" in daily_df.columns:
        code_str = (
            daily_df["代码"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
            .str.zfill(6)
        )
        code_to_name = dict(zip(code_str, daily_df["名称"].astype(object)))
    existing = pd.read_excel(path, sheet_name="每日资金流入汇总")
    # 旧表中的股票列也统一为 6 位字符串，避免 merge 时出现 int64 vs str
    if "股票" in existing.columns:
        existing["股票"] = (
            existing["股票"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
            .str.zfill(6)
        )
    for c in add_cols:
        if c != "股票" and c in existing.columns:
            existing = existing.drop(columns=[c])
    # outer：既有股票更新当日列，新股追加新行（上市前日期列为空）
    merged = existing.merge(add[add_cols], on="股票", how="outer")
    new_count = len(merged) - len(existing)
    if new_count > 0:
        print(f"  每日宽表新增 {new_count} 只新股（上市前日期列为空，可后续用 fill_daily_history 按需补全）")
    if code_to_name and "股票名称" in merged.columns:
        merged["股票名称"] = merged["股票名称"].fillna(merged["股票"].map(code_to_name))
    # 涨跌幅列保持 object，避免写入字符串时报错
    for c in merged.columns:
        if "涨跌幅" in str(c):
            merged[c] = merged[c].astype(object)
    write_daily_wide_workbook(daily_wide_df=merged, output_path=path)
